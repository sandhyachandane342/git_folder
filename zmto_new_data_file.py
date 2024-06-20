import psycopg2
from openpyxl import Workbook, load_workbook
conn = psycopg2.connect(
    host='borosilcp.postgres.database.azure.com',
    database='easyecom',
    user='db_user',
    password='Hello123!'
    )

def get_reference_codes_from_file(file_path):
  reference_codes = []
  workbook = load_workbook(file_path)
  worksheet = workbook.active
  for row in worksheet.iter_rows(values_only = True):
    #print(row)
    reference_codes.append(str(row[0]))
  return reference_codes
cursor = conn.cursor()
file_path = input("Enter the file path: ")
reference_codes = get_reference_codes_from_file(file_path)
reference_codes = tuple(reference_codes)
print(reference_codes)

sql_query = '''
with sub as (
    select 'ZMTO' "Document Type", 
           concat(reference_code, '_', suborder_num) "Order Num_Sub Order Num", 
           order_date "Order Date", 
           case when invoice_number is null then '--' else invoice_number end as "Invoice Number", 
           manifest_date "Invoice Date", 
           sku "SKU", 
           item_quantity "Quantity",
           case when breakup_types ilike '%%IGST%%' then 'IGST' else 'SGST' end as "Tax Type",
           tax_rate "Percentage", 
           'EC0025' "Marketplace", 
           '7200' "Sales Organization",
           'EC' "Distribution Channel Code", 
           '0' "Division Code",
           case when warehouse_id = 96778 then 7207 else null end as "Sales Office Code",
           case when warehouse_id = 96778 then 1231 else null end as "Plant Code",
           'VMYB' "Storage Location Code", 
           billing_name "Bill To Name",
           concat(billing_address, ',', billing_city, '-', billing_pin_code, ',', billing_state) "Bill To Address", 
           billing_mobile "Bill To Contact", 
           billing_pin_code "Bill To Pincode", 
           billing_state "Bill To City", 
           null "Bill To Region", 
           shipping_name "Ship To Name", 
           concat(shipping_address, ',', shipping_city, '-', shipping_pin_code, ',', shipping_state) "Ship To Address", 
           shipping_mobile "Ship To Contact", 
           shipping_pin_code "Ship To Pincode", 
           shipping_state "Ship To City", 
           null "Ship To Region", 
           'West' "Zone", 
           36 "RegionCode",
           state_customer.customer_code "Bill to party", 
           state_customer.customer_code "Payer",
           state_customer.customer_code "Sold to party", 
           state_customer.customer_code "Ship to party",
           case when buyer_gst = 'NA' then null else buyer_gst end as "GSTN", 
           1 "Exchange Rate", 
           'INR' "Currency",
           case when breakup_types ilike '%%IGST%%' 
                then round(((breakup_types::jsonb->>'Item Amount Excluding Tax')::numeric + 
                            (breakup_types::jsonb->>'Item Amount IGST')::numeric) / item_quantity)
                else round(((breakup_types::jsonb->>'Item Amount Excluding Tax')::numeric + 
                            (breakup_types::jsonb->>'Item Amount CGST')::numeric + 
                            (breakup_types::jsonb->>'Item Amount SGST')::numeric) / item_quantity)
           end as "ZMRP",
           case when breakup_types ilike '%%IGST%%' 
                then round(((breakup_types::jsonb->>'Shipping Excluding Tax')::numeric + 
                            (breakup_types::jsonb->>'Shipping IGST')::numeric) / item_quantity, 2)
                else round(((breakup_types::jsonb->>'Shipping Excluding Tax')::numeric + 
                            (breakup_types::jsonb->>'Shipping CGST')::numeric + 
                            (breakup_types::jsonb->>'Shipping SGST')::numeric) / item_quantity, 2)
           end as "ZDEL",
           case when breakup_types ilike '%%IGST%%' 
                then round(((breakup_types::jsonb->>'Promotion Discount Excluding Tax')::numeric + 
                            (breakup_types::jsonb->>'Promotion Discount IGST')::numeric) / item_quantity, 2)
                else round(((breakup_types::jsonb->>'Promotion Discount Excluding Tax')::numeric + 
                            (breakup_types::jsonb->>'Promotion Discount CGST')::numeric + 
                            (breakup_types::jsonb->>'Promotion Discount SGST')::numeric) / item_quantity, 2)
           end as "ZED1",
           (select data->>'field_value'
            from jsonb_array_elements(custom_fields::jsonb) data
            where data->>'field_name' = 'Coupon Code' or data->>'field_name' = 'gift_card_number' or 
                  data->>'field_name' = 'qc_card_numbers') "COUPON CODE NO",
           coalesce(breakup_types::jsonb->>'PromoCode Discount Excluding Tax', 
                    (select data->>'field_value'
                     from jsonb_array_elements(custom_fields::jsonb) data
                     where data->>'field_name' = 'gift_card_redeemed_amount')) "ZVOU",
           awb_number "AWB NO", 
           manifest_date "AWB Date",
           case when custom_fields like '%%gift_card_number%%' then 'R' 
                when custom_fields like '%%qc_card_numbers%%' then 'P'	 
                else null end as "Gift Voucher Flag",
           payment_mode "Transaction Data", 
           invoice_id "EasyEcomInvoiceID", 
           manifest_timestamp "Manifest Date/Time",
           shipping_state_code "Customer State Code", 
           (invoice_documents::jsonb)->>'irn' "IRN Number", 
           (invoice_documents::jsonb)->>'AckNo' "Acknowledgment Number", 
           null "IRN E-Way Bill Number",
           (select data->>'field_value'
            from jsonb_array_elements(custom_fields::jsonb) data
            where data->>'field_name' = 'Placement') "Placement",
           (select data->>'field_value'
            from jsonb_array_elements(custom_fields::jsonb) data
            where data->>'field_name' = 'Name') "Name",
           (select data->>'field_value'
            from jsonb_array_elements(custom_fields::jsonb) data
            where data->>'field_name' = 'Font') "Font",
           (select data->>'field_value'
            from jsonb_array_elements(custom_fields::jsonb) data
            where data->>'field_name' = 'Customise-charge') "Customise-charge",
           (select data->>'field_value'
            from jsonb_array_elements(custom_fields::jsonb) data
            where data->>'field_name' = 'CustomImage') "CustomImage"
    from easyecom_suborders
    left join state_customer 
        on state_customer.state ilike easyecom_suborders.billing_state and 
           state_customer.flag = (select case when easyecom_suborders.warehouse_id = 32682 then 'jaipur' else 'alt' end as flag)
    where warehouse_id = 96778 AND reference_code IN %s
)
select sub."Document Type",
       sub."Order Num_Sub Order Num", 
       sub."Order Date", 
       sub."Invoice Number", 
       sub."Invoice Date", 
       sub."SKU", 
       sub."Quantity", 
       coalesce(master_sku.accounting_unit, 'EA') as "Unit",  -- Use COALESCE to replace null with 'EA'
       sub."Tax Type", 
       sub."Percentage", 
       sub."Marketplace", 
       sub."Sales Organization", 
       sub."Distribution Channel Code", 
       sub."Division Code", 
       sub."Sales Office Code", 
       sub."Plant Code", 
       sub."Storage Location Code", 
       sub."Bill To Name", 
       sub."Bill To Address", 
       sub."Bill To Contact", 
       sub."Bill To Pincode", 
       sub."Bill To City", 
       sub."Bill To Region", 
       sub."Ship To Name", 
       sub."Ship To Address", 
       sub."Ship To Contact", 
       sub."Ship To Pincode", 
       sub."Ship To City", 
       sub."Ship To Region", 
       sub."Zone", 
       sub."RegionCode", 
       sub."Bill to party", 
       sub."Payer", 
       sub."Sold to party", 
       sub."Ship to party", 
       sub."GSTN", 
       sub."Exchange Rate", 
       sub."Currency", 
       sub."ZMRP", 
       sub."ZDEL", 
       sub."ZED1", 
       sub."COUPON CODE NO", 
       sub."ZVOU", 
       sub."AWB NO", 
       sub."AWB Date", 
       sub."Gift Voucher Flag", 
       sub."Transaction Data", 
       sub."EasyEcomInvoiceID", 
       sub."Manifest Date/Time", 
       sub."Customer State Code", 
       sub."IRN Number", 
       sub."Acknowledgment Number", 
       sub."IRN E-Way Bill Number",
       concat('|Placement:', sub."Placement", '|Name:', sub."Name", '|Font:', sub."Font", '|Customise-charge:', sub."Customise-charge", '|CustomImage:', sub."CustomImage") "Personalization", 
       zecomm_new.sales_ord_no as "Sales Order"
from sub
left join master_sku on master_sku.sku = sub."SKU"
left join zecomm_new on sub."Order Num_Sub Order Num" = zecomm_new.fba_ord_no;
'''
cursor.execute(sql_query, (reference_codes,))
data = cursor.fetchall()
#print(data)

cursor.close()
conn.close()

workbook = Workbook()
worksheet = workbook.active
headers = [desc[0] for desc in cursor.description]
print(headers)

for col_index, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col_index, value=header)

for row_index, row in enumerate(data, start=2):
    for col_index, value in enumerate(row, start=1):
         worksheet.cell(row=row_index, column=col_index, value=value)

file_name = "easyecom_data_new.xlsx"
workbook.save(file_name)
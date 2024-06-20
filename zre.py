import psycopg2
from openpyxl import Workbook, load_workbook

conn = psycopg2.connect(
    host='borosilcp.postgres.database.azure.com',
    database='easyecom',
    user='db_user',
    password='Hello123!'
)

def get_reference_codes_from_file(file_path):
    reference_codes = []
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    for row in worksheet.iter_rows(values_only=True):
        reference_codes.append(str(row[0]))  
    return reference_codes

cursor = conn.cursor()

file_path = input("Enter the file path : ")

reference_codes = get_reference_codes_from_file(file_path)
print(reference_codes)

sql_query = '''select
	'ZRE' as "Document Type",
	sap.invoice_no as "Sales Order Id",
	suborder->>'sku' as "Material",
	suborder->>'returned_item_quantity' as "Quantity",
	ms.accounting_unit as "Unit of Measure",
	ee.return_date as "Return Date",
	sap.plant as "Plant to",
	'RMBO' as "Storage Location to",
	453 as "Type",
	concat(reference_code, '_', credit_note_id) as "Order Id",
	ee.return_date as "Return Date",
	ee.credit_note_number as "CreditNote_Num"
from all_returns ee
cross join lateral jsonb_array_elements(ee.items) suborder
left join zecomm_new sap on
	(case when "warehouseId" <> 96778 then concat(ee.reference_code, '_', ee.invoice_number)
	else concat(ee.reference_code, '_', suborder->>'suborder_num') end) = sap.fba_ord_no and
	suborder->>'sku' = sap.material_code
left join master_sku ms on suborder->>'sku' = ms.sku
WHERE reference_code IN %s
'''
    
cursor.execute(sql_query, (tuple(reference_codes),))
data = cursor.fetchall()
print(data)

workbook = Workbook()
worksheet = workbook.active
headers = [desc[0] for desc in cursor.description]

for col_index, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col_index, value=header)

for row_index, row in enumerate(data, start=2):
    for col_index, value in enumerate(row, start=1):
         worksheet.cell(row=row_index, column=col_index, value=value)

file_name = "data_new.xlsx"
workbook.save(file_name)
    
cursor.close()
conn.close()

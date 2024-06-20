import psycopg2
from openpyxl import Workbook, load_workbook
conn = psycopg2.connect(
    host='borosilcp.postgres.database.azure.com',
    database='easyecom',
    user='db_user',
    password='Hello123!'
)
def get_reference_codes_from_file(file_path):
  reference_codes = []
  workbook = load_workbook(file_path)
  worksheet = workbook.active
  for row in worksheet.iter_rows(values_only = True):
    #print(row)
    reference_codes.append(str(row[0]))
  return reference_codes
cursor = conn.cursor()
file_path = input("Enter the file path: ")
reference_codes = get_reference_codes_from_file(file_path)
reference_codes = tuple(reference_codes)
print(reference_codes)
sql_query = '''
with order_data AS(
  select 'ZEC' "Document Type", 
		case
			when marketplace_id in (26,517,104,2) then concat(reference_code, '_', invoice_number)
			when marketplace_id = 233 then concat(suborder->>'invoicecode', '_', invoice_number)
		end as "Order Num",
		order_date::date "Order Date", invoice_number as "Invoice Number", 
		case when invoice_date = '' then null
		else invoice_date::date end as "Invoice Date", split_part(suborder->>'sku', '-', 1) "SKU", 
		suborder->>'item_quantity' "Quantity",
		case when suborder->>'breakup_types' ilike '%%IGST%%' then 'IGST'
		else 'SGST' end as "Tax Type",
		suborder->>'tax_rate' "Percentage", 'EC0025' "Marketplace", '7200' "Sales Organization",
		'EC' "Distribution Channel Code", '0' "Division Code",
		case when assigned_warehouse_id in (61827,14436) then 7202
		when assigned_warehouse_id = 48305 then 7203
		when assigned_warehouse_id in (32682,61866) then 7201 --Jaipur
		when assigned_warehouse_id in (82835,85219) then 7205 --Bangalore
		when assigned_warehouse_id = 86051 then 7204
		when assigned_warehouse_id = 124049 then 7202
		else null end as "Sales Office Code",
		case when assigned_warehouse_id in (61827,14436) then 1269
		when assigned_warehouse_id = 48305 then 1284
		when assigned_warehouse_id in (32682,61866) then 7201 --Jaipur
		when assigned_warehouse_id in (82835,85219) then 1285 --Bangalore
		when assigned_warehouse_id = 86051 then 1286
		when assigned_warehouse_id = 124049 then 1269
		else null end as "Plant Code",
		'VMYB' "Storage Location Code", billing_name "Bill To Name",
		billing_address_1 "Bill To Address", billing_mobile "Bill To Contact", 
		billing_pin_code "Bill To Pincode", billing_state "Bill To City", null "Bill To Region", 
		shipping_name "Ship To Name", address_line_1 "Ship To Address", 
		contact_num "Ship To Contact", pin_code "Ship To Pincode", order_details.state "Ship To City", 
		null "Ship To Region", 'West' "Zone", 36 "RegionCode",
		state_customer.customer_code "Bill to party", state_customer.customer_code "Payer",
		state_customer.customer_code "Sold to party", state_customer.customer_code "Ship to party",
		case when buyer_gst = 'NA' then null else buyer_gst end as "GSTN", 
		1 "Exchange Rate", 'INR' "Currency",
		case when suborder->>'breakup_types' ilike '%%IGST%%' 
			 then round(((suborder->'breakup_types'->>'Item Amount Excluding Tax')::numeric + 
			 	  (suborder->'breakup_types'->>'Item Amount IGST')::numeric)/(suborder->>'item_quantity')::integer)
		else round(((suborder->'breakup_types'->>'Item Amount Excluding Tax')::numeric +
		 	 (suborder->'breakup_types'->>'Item Amount CGST')::numeric +
		 	 (suborder->'breakup_types'->>'Item Amount SGST')::numeric)/(suborder->>'item_quantity')::integer)
		end as "ZMRP",
		case when suborder->>'breakup_types' ilike '%%IGST%%' 
			 then round(((suborder->'breakup_types'->>'Shipping Excluding Tax')::numeric + 
			 	  (suborder->'breakup_types'->>'Shipping IGST')::numeric)/(suborder->>'item_quantity')::integer,2)
		else round(((suborder->'breakup_types'->>'Shipping Excluding Tax')::numeric +
		 	 (suborder->'breakup_types'->>'Shipping CGST')::numeric +
		 	 (suborder->'breakup_types'->>'Shipping SGST')::numeric)/(suborder->>'item_quantity')::integer,2)
		end as "ZDEL",
		case when suborder->>'breakup_types' ilike '%%IGST%%' 
			 then round(((suborder->'breakup_types'->>'Promotion Discount Excluding Tax')::numeric + 
			 	  (suborder->'breakup_types'->>'Promotion Discount IGST')::numeric)/(suborder->>'item_quantity')::integer,2)
		else round(((suborder->'breakup_types'->>'Promotion Discount Excluding Tax')::numeric +
		 	 (suborder->'breakup_types'->>'Promotion Discount CGST')::numeric +
		 	 (suborder->'breakup_types'->>'Promotion Discount SGST')::numeric)/(suborder->>'item_quantity')::integer,2)
		end as "ZED1",
		(select data->>'field_value'
        FROM jsonb_array_elements(suborder->'custom_fields') data
        WHERE data->>'field_name' = 'Coupon Code' or data->>'field_name' = 'gift_card_number' or 
        data->>'field_name' = 'qc_card_numbers') "COUPON CODE NO",
		coalesce(suborder->'breakup_types'->>'PromoCode Discount Excluding Tax', (select data->>'field_value'
																		FROM jsonb_array_elements(suborder->'custom_fields') data
        WHERE data->>'field_name' = 'gift_card_redeemed_amount')) "ZVOU",
		awb_number "AWB NO", manifest_date::date "AWB Date",
		case when suborder->>'custom_fields' like '%%gift_card_number%%' then 'R' 
		when suborder->>'custom_fields' like '%%qc_card_numbers%%' then 'P'	 
		else null end as "Gift Voucher Flag",
		payment_mode "Transaction Data", invoice_id "EasyEcomInvoiceID", manifest_date "Manifest Date/Time",
		state_code "Customer State Code", (invoice_documents::jsonb)->>'irn' "IRN Number", 
		(invoice_documents::jsonb)->>'AckNo' "Acknowledgment Number", 
		null "IRN E-Way Bill Number"
from order_details
left join state_customer 
	on state_customer.state ilike order_details.billing_state and 
	state_customer.flag = (select case when order_details.assigned_warehouse_id = 32682 then 'jaipur'
													   else 'alt' end as flag)
cross join lateral jsonb_array_elements(order_details.order_items) suborder
where assigned_warehouse_id <> 96778 AND reference_code IN %s
)
select "Document Type", "Order Num", "Order Date", "Invoice Number", "Invoice Date", "SKU", "Quantity", master_sku.accounting_unit "Unit", "Tax Type", "Percentage", "Marketplace", "Sales Organization", "Distribution Channel Code", "Division Code", "Sales Office Code", "Plant Code", "Storage Location Code", "Bill To Name", "Bill To Address", "Bill To Contact", "Bill To Pincode", "Bill To City", "Bill To Region", "Ship To Name", "Ship To Address", "Ship To Contact", "Ship To Pincode", "Ship To City", "Ship To Region", "Zone", "RegionCode", "Bill to party", "Payer", "Sold to party", "Ship to party", "GSTN", "Exchange Rate", "Currency", "ZMRP", "ZDEL", "ZED1", "COUPON CODE NO", "ZVOU", "AWB NO", "AWB Date", "Gift Voucher Flag", "Transaction Data", "EasyEcomInvoiceID", "Manifest Date/Time", "Customer State Code", "IRN Number", "Acknowledgment Number", "IRN E-Way Bill Number", zecomm_new.sales_ord_no AS "Sales Order"
FROM order_data
LEFT JOIN
master_sku on master_sku.sku = order_data."SKU"
LEFT JOIN zecomm_new
ON zecomm_new.fba_ord_no = order_data."Order Num";
'''
cursor.execute(sql_query, (reference_codes,))
data = cursor.fetchall()
#print(data)

workbook = Workbook()
worksheet = workbook.active
headers = [desc[0] for desc in cursor.description]
print(headers)

for col_index, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col_index, value=header)

for row_index, row in enumerate(data, start=2):
    for col_index, value in enumerate(row, start=1):
         worksheet.cell(row=row_index, column=col_index, value=value)

file_name = "data_new.xlsx"
workbook.save(file_name)

cursor.close()
conn.close()